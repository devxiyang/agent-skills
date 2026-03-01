#!/usr/bin/env python3
"""
Generate a pivot table from an Excel or CSV file.

Usage:
    python3 excel_pivot.py data.xlsx --index Region --columns Product --values Revenue
    python3 excel_pivot.py data.csv --index Category --values Amount --agg sum
    python3 excel_pivot.py data.xlsx --index "Region,Year" --values "Revenue,Cost" --agg sum

Arguments:
    --index    Column(s) to use as row labels (comma-separated for multiple)
    --columns  Column to use as column labels (optional)
    --values   Column(s) to aggregate (comma-separated for multiple)
    --agg      Aggregation function: sum, mean, count, min, max (default: sum)
    --output   Output file path (default: <input>_pivot.xlsx)
"""

import argparse
import sys
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print('Missing dependencies. Run: pip install pandas openpyxl')
    sys.exit(1)


def read_input(path: Path) -> 'pd.DataFrame':
    if path.suffix.lower() == '.csv':
        return pd.read_csv(path)
    return pd.read_excel(path)


def style_pivot_sheet(ws) -> None:
    header_fill = PatternFill('solid', fgColor='1E3A5F')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    index_fill  = PatternFill('solid', fgColor='DBEAFE')
    total_fill  = PatternFill('solid', fgColor='FEF9C3')
    total_font  = Font(bold=True)

    max_row = ws.max_row
    max_col = ws.max_column

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Header row
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill

            # Index column (first column)
            elif cell.column == 1 and cell.row > 1:
                cell.font = Font(bold=False)
                cell.fill = index_fill

            # Total row (last row if margins=True)
            if str(cell.value) in ('Total', 'All'):
                cell.font = total_font
                cell.fill = total_fill

    # Auto column width
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    ws.freeze_panes = 'B2'


def main() -> None:
    parser = argparse.ArgumentParser(description='Generate pivot table from Excel/CSV')
    parser.add_argument('input', help='Input .xlsx or .csv file')
    parser.add_argument('--index', required=True, help='Row label column(s), comma-separated')
    parser.add_argument('--columns', default=None, help='Column label column (optional)')
    parser.add_argument('--values', required=True, help='Value column(s), comma-separated')
    parser.add_argument('--agg', default='sum', choices=['sum','mean','count','min','max'],
                        help='Aggregation function (default: sum)')
    parser.add_argument('--output', '-o', help='Output .xlsx file')
    parser.add_argument('--no-totals', action='store_true', help='Skip totals row/column')
    args = parser.parse_args()

    src = Path(args.input)
    if not src.exists():
        print(f'Error: {src} not found')
        sys.exit(1)

    out = Path(args.output) if args.output else src.with_name(src.stem + '_pivot.xlsx')

    index  = [c.strip() for c in args.index.split(',')]
    values = [c.strip() for c in args.values.split(',')]
    cols   = args.columns.strip() if args.columns else None

    print(f'Reading {src}...')
    df = read_input(src)
    print(f'  {len(df)} rows × {len(df.columns)} columns')

    print(f'Building pivot: index={index}, columns={cols}, values={values}, agg={args.agg}')
    pivot = df.pivot_table(
        index=index,
        columns=cols,
        values=values if len(values) > 1 else values[0],
        aggfunc=args.agg,
        fill_value=0,
        margins=not args.no_totals,
        margins_name='Total',
    ).round(2)

    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        pivot.to_excel(writer, sheet_name='Pivot')
        df.to_excel(writer, sheet_name='Source', index=False)

    wb = load_workbook(out)
    style_pivot_sheet(wb['Pivot'])
    wb.save(out)

    print(f'Saved → {out}')
    print(f'  Pivot shape: {pivot.shape[0]} rows × {pivot.shape[1]} columns')


if __name__ == '__main__':
    main()
