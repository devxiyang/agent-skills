#!/usr/bin/env python3
"""
Generate a styled Excel report from a CSV file.

Usage:
    python3 excel_report.py data.csv
    python3 excel_report.py data.csv --output report.xlsx --title "Q3 Sales Report"

Input CSV must have a header row. All columns will be included.
A summary sheet with column stats is generated automatically.
"""

import argparse
import sys
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.chart import BarChart, Reference
except ImportError:
    print('Missing dependencies. Run: pip install pandas openpyxl')
    sys.exit(1)


HEADER_BG   = '2563EB'
HEADER_FG   = 'FFFFFF'
ACCENT_BG   = 'DBEAFE'
ALT_ROW_BG  = 'F8FAFC'


def style_header(ws, n_cols: int) -> None:
    thin = Side(style='thin', color='CBD5E1')
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color=HEADER_FG, size=11)
        cell.fill = PatternFill('solid', fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(bottom=Side(style='medium', color='1D4ED8'))
    ws.row_dimensions[1].height = 28


def style_data_rows(ws, n_rows: int, n_cols: int) -> None:
    thin = Side(style='thin', color='E2E8F0')
    for row in range(2, n_rows + 2):
        bg = ALT_ROW_BG if row % 2 == 0 else 'FFFFFF'
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill('solid', fgColor=bg)
            cell.border = Border(
                bottom=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0'),
            )
            cell.alignment = Alignment(vertical='center')


def auto_column_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def add_table(ws, n_rows: int, n_cols: int, name: str = 'DataTable') -> None:
    ref = f'A1:{get_column_letter(n_cols)}{n_rows + 1}'
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2',
        showRowStripes=True,
    )
    ws.add_table(table)


def build_summary(df: 'pd.DataFrame') -> 'pd.DataFrame':
    numeric = df.select_dtypes(include='number')
    if numeric.empty:
        return pd.DataFrame({'Note': ['No numeric columns found']})
    return numeric.describe().T.rename(columns={
        'count': 'Count', 'mean': 'Average', 'std': 'Std Dev',
        'min': 'Min', '25%': 'Q1', '50%': 'Median', '75%': 'Q3', 'max': 'Max',
    }).round(2).reset_index().rename(columns={'index': 'Column'})


def add_bar_chart(ws, df: 'pd.DataFrame') -> None:
    numeric_cols = df.select_dtypes(include='number').columns
    if len(numeric_cols) == 0:
        return
    n_rows = min(len(df), 20)
    n_cols = min(len(numeric_cols), 4)

    chart = BarChart()
    chart.type = 'col'
    chart.title = 'Data Overview'
    chart.style = 10
    chart.y_axis.title = 'Value'
    chart.width = 22
    chart.height = 12

    data = Reference(ws, min_col=2, max_col=n_cols + 1, min_row=1, max_row=n_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1))

    col_after = get_column_letter(len(df.columns) + 2)
    ws.add_chart(chart, f'{col_after}2')


def main() -> None:
    parser = argparse.ArgumentParser(description='Generate a styled Excel report from CSV')
    parser.add_argument('input', help='Input CSV file')
    parser.add_argument('--output', '-o', help='Output .xlsx file (default: <input>.xlsx)')
    parser.add_argument('--title', '-t', default='Report', help='Report title (shown in sheet tab)')
    parser.add_argument('--no-chart', action='store_true', help='Skip chart generation')
    args = parser.parse_args()

    src = Path(args.input)
    if not src.exists():
        print(f'Error: {src} not found')
        sys.exit(1)

    out = Path(args.output) if args.output else src.with_suffix('.xlsx')

    print(f'Reading {src}...')
    df = pd.read_csv(src)
    print(f'  {len(df)} rows × {len(df.columns)} columns')

    summary_df = build_summary(df)

    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=args.title[:31], index=False)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

    wb = load_workbook(out)

    # Style data sheet
    ws = wb[args.title[:31]]
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    style_header(ws, len(df.columns))
    style_data_rows(ws, len(df), len(df.columns))
    auto_column_width(ws)
    if not args.no_chart:
        add_bar_chart(ws, df)

    # Style summary sheet
    ws_sum = wb['Summary']
    ws_sum.freeze_panes = 'A2'
    style_header(ws_sum, len(summary_df.columns))
    style_data_rows(ws_sum, len(summary_df), len(summary_df.columns))
    auto_column_width(ws_sum)

    wb.save(out)
    print(f'Saved → {out}')


if __name__ == '__main__':
    main()
